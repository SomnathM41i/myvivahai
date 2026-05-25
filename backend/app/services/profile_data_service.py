"""
app/services/profile_data_service.py
Business logic for generated profile/biodata management.
Handles view, edit, delete, and multi-format export.
"""
import csv
import io
import json
from datetime import datetime
from typing import Any

from fastapi import HTTPException
from sqlalchemy.ext.asyncio import AsyncSession

from app.repositories.profile_repository import ProfileRepository
from app.repositories.upload_repository import UploadRepository
from app.schemas.profile_schema import ProfileSchema, ProfileUpdateSchema


# Human-readable labels for export columns
_EXPORT_LABELS: dict[str, str] = {
    "full_name":          "Full Name",
    "date_of_birth":      "Date of Birth",
    "gender":             "Gender",
    "religion":           "Religion",
    "caste":              "Caste",
    "mother_tongue":      "Mother Tongue",
    "height":             "Height",
    "blood_group":        "Blood Group",
    "mobile":             "Mobile",
    "email":              "Email",
    "city":               "City",
    "state":              "State",
    "education":          "Education",
    "occupation":         "Occupation",
    "annual_income":      "Annual Income",
    "father_name":        "Father's Name",
    "mother_name":        "Mother's Name",
    "siblings":           "Siblings",
    "family_type":        "Family Type",
    "rashi":              "Rashi",
    "nakshatra":          "Nakshatra",
    "gotra":              "Gotra",
    "manglik":            "Manglik",
    "partner_preference": "Partner Preference",
    "ai_confidence":      "AI Confidence",
}


class ProfileDataService:
    def __init__(self, db: AsyncSession):
        self._db = db
        self._profile_repo = ProfileRepository(db)
        self._upload_repo = UploadRepository(db)

    async def get_profile(self, profile_id: int, user_id: int):
        profile = await self._get_owned(profile_id, user_id)
        return profile

    async def update_profile(
        self,
        profile_id: int,
        user_id: int,
        data: ProfileUpdateSchema,
    ):
        profile = await self._get_owned(profile_id, user_id)
        updates = data.model_dump(exclude_none=True)
        if not updates:
            return profile

        profile = await self._profile_repo.update(profile, **updates)

        # Also keep upload's processed_output in sync if present
        if profile.upload_id:
            upload = await self._upload_repo.get_by_id(profile.upload_id)
            if upload and upload.processed_output:
                try:
                    raw = json.loads(upload.processed_output)
                    raw.update(updates)
                    await self._upload_repo.update_status(
                        upload, upload.status, processed_output=json.dumps(raw)
                    )
                except Exception:
                    pass  # Don't block save if sync fails

        await self._db.commit()
        return profile

    async def delete_profile(self, profile_id: int, user_id: int) -> None:
        profile = await self._get_owned(profile_id, user_id)
        await self._profile_repo.delete(profile)
        await self._db.commit()

    async def export_json(self, profile_id: int, user_id: int) -> bytes:
        profile = await self._get_owned(profile_id, user_id)
        data = self._profile_to_export_dict(profile)
        data["exported_at"] = datetime.utcnow().isoformat()
        return json.dumps(data, indent=2, ensure_ascii=False).encode("utf-8")

    async def export_csv(self, profile_id: int, user_id: int) -> bytes:
        profile = await self._get_owned(profile_id, user_id)
        data = self._profile_to_export_dict(profile)

        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(["Field", "Value"])
        for label, value in data.items():
            writer.writerow([label, value])
        writer.writerow(["Exported At", datetime.utcnow().isoformat()])
        return buf.getvalue().encode("utf-8")

    async def export_xlsx(self, profile_id: int, user_id: int) -> bytes:
        """
        Build an xlsx file using openpyxl (already in your deps for OCR pipeline).
        Returns raw bytes.
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise HTTPException(500, "openpyxl not installed — cannot generate xlsx")

        profile = await self._get_owned(profile_id, user_id)
        raw_data = self._profile_to_export_dict(profile)
        # Eagerly convert all values to plain Python strings while the
        # SQLAlchemy session is still open — prevents DetachedInstanceError
        # and openpyxl serialisation failures on lazy-loaded attributes.
        data: dict = {
            str(label): str(value) if value is not None else ""
            for label, value in raw_data.items()
        }
        data["Exported At"] = datetime.utcnow().isoformat()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Profile"

        # Header row
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4F46E5")  # indigo-600
        for col, heading in enumerate(["Field", "Value"], start=1):
            cell = ws.cell(row=1, column=col, value=heading)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for row_idx, (label, value) in enumerate(data.items(), start=2):
            ws.cell(row=row_idx, column=1, value=label)
            ws.cell(row=row_idx, column=2, value=value)

        # Column widths
        ws.column_dimensions["A"].width = 24
        ws.column_dimensions["B"].width = 42

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # ── private ────────────────────────────────────────────────────────────
    async def _get_owned(self, profile_id: int, user_id: int):
        profile = await self._profile_repo.get_by_id(profile_id)
        if not profile or profile.user_id != user_id:
            raise HTTPException(404, "Profile not found")
        return profile

    def _profile_to_export_dict(self, profile) -> dict[str, Any]:
        result: dict[str, Any] = {}
        for field, label in _EXPORT_LABELS.items():
            value = getattr(profile, field, None)
            if value is not None and value != "":
                if field == "ai_confidence":
                    result[label] = f"{round(float(value) * 100)}%"
                else:
                    result[label] = value
        return result